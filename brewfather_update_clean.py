#!/usr/bin/env python3
r"""
Brewfather master batch and recipe updater and clean export generator.

Each run:
  1. Downloads every Brewfather batch with complete=True and every recipe.
  2. Loads the existing local master JSONs, if present.
  3. Upserts batches and recipes by Brewfather _id:
       - new items are added;
       - changed items replace the older local version;
       - historical local items missing from the API response are retained.
  4. Regenerates clean TXT, Markdown and Excel outputs from the merged masters.

Files maintained in --out-dir:
  Brewfather_Batches_Master.json + _Readable_Clean.{txt,md}, _Summary.xlsx
  Brewfather_Recipes_Master.json + _Readable_Clean.{txt,md}, _Summary.xlsx
  Brewfather_Runs_Log.csv
      Append-only audit trail: one row per run with batch/recipe counts,
      master SHA-256 hashes, backup references and changed IDs.
      Disable with --no-run-log.
  history/Brewfather_Batches_Master_YYYYMMDD_HHMMSS.json
  history/Brewfather_Recipes_Master_YYYYMMDD_HHMMSS.json
      Created automatically before a changed master is overwritten.

Credentials are read from environment variables by default:
  BREWFATHER_USER_ID
  BREWFATHER_API_KEY

Windows PowerShell example:
  pip install requests openpyxl
  $env:BREWFATHER_USER_ID="your_user_id"
  $env:BREWFATHER_API_KEY="your_api_key"
  python brewfather_update_clean.py --out-dir "C:\pycode\beer\exports"

The script fetches all batches on each run because this also refreshes later changes
such as measured FG, packaging date, status, notes and tasting results. The on-disk
master is updated incrementally and is never pruned unless the code is deliberately
changed.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import shutil
import sys
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import requests
except ImportError:  # handled in main
    requests = None

try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None

API_BASE_URL = "https://api.brewfather.app/v2"
MAX_LIMIT = 50
SCRIPT_VERSION = "2.1.0"


# -----------------------------------------------------------------------------
# Basic helpers
# -----------------------------------------------------------------------------


def get_nested(obj: Dict[str, Any], *keys: str, default: Any = "") -> Any:
    cur: Any = obj
    for key in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(key, default)
    return default if cur is None else cur


def first_value(*values: Any, default: Any = "") -> Any:
    for value in values:
        if value is not None and value != "":
            return value
    return default


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def fmt_number(value: Any, decimals: int = 2, blank: str = "") -> str:
    if value is None or value == "":
        return blank
    if isinstance(value, str):
        return value
    if not is_number(value):
        return str(value)
    text = f"{float(value):.{decimals}f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def fmt_gravity(value: Any, blank: str = "") -> str:
    if value is None or value == "":
        return blank
    try:
        return f"{float(value):.3f}"
    except (TypeError, ValueError):
        return str(value)

def fmt_amount(value: Any, unit: str = "", decimals: int = 3) -> str:
    if value is None or value == "":
        return ""
    text = fmt_number(value, decimals=decimals)
    return f"{text} {unit}".strip()


def timestamp_to_datetime(value: Any, tz_name: str = "Africa/Johannesburg") -> Optional[datetime]:
    """Convert Brewfather millisecond timestamps to timezone-aware datetimes."""
    if value is None or value == "":
        return None

    # Firestore-style timestamp object
    if isinstance(value, dict):
        value = value.get("_seconds")
    try:
        seconds = float(value)
    except (TypeError, ValueError):
        return None
    # Brewfather date fields are normally milliseconds.
    if seconds > 10_000_000_000:
        seconds = seconds / 1000.0
    # 0 or negative values are missing data, not the 1970 epoch.
    if seconds <= 0:
        return None

    try:
        if ZoneInfo is not None:
            tz = ZoneInfo(tz_name)
        else:
            tz = timezone(timedelta(hours=2))
        return datetime.fromtimestamp(seconds, tz=tz)
    except Exception:
        return None


def fmt_date(value: Any, tz_name: str = "Africa/Johannesburg", with_time: bool = False) -> str:
    dt = timestamp_to_datetime(value, tz_name=tz_name)
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M") if with_time else dt.strftime("%Y-%m-%d")


def now_in_tz(tz_name: str) -> datetime:
    """Current time in the configured timezone, so run timestamps match batch dates."""
    if ZoneInfo is not None:
        try:
            return datetime.now(ZoneInfo(tz_name))
        except Exception:
            pass
    return datetime.now(timezone(timedelta(hours=2)))


def brew_date_sort_key(batch: Dict[str, Any]) -> float:
    """Numeric sort key tolerant of missing, string or Firestore-dict brew dates."""
    dt = timestamp_to_datetime(batch.get("brewDate"))
    return dt.timestamp() if dt is not None else 0.0


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def section(lines: List[str], title: str, markdown: bool = False, underline: str = "-") -> None:
    if markdown:
        # Explicit heading level below the "## Batch ..." title; the previous
        # dash underline rendered as a setext H2 at the same level as the title.
        lines.append(f"### {title}")
    else:
        lines.append(title)
        lines.append(underline * 20)


def add_multiline(lines: List[str], value: Any) -> None:
    text = clean_text(value)
    if text:
        lines.extend(text.splitlines())
    else:
        lines.append("")


def safe_list(value: Any) -> List[Dict[str, Any]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    return []


def canonical_json(value: Any) -> str:
    """Stable JSON representation used to identify changed batch records."""
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def without_api_readings(batch: Dict[str, Any]) -> Dict[str, Any]:
    """Return a shallow copy without locally attached device readings."""
    if "apiReadings" not in batch:
        return batch
    clean = dict(batch)
    clean.pop("apiReadings", None)
    return clean


def load_batch_list(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise RuntimeError(f"Expected a JSON list in {path}")
    return [item for item in data if isinstance(item, dict)]


def write_json_atomic(path: Path, data: Any) -> None:
    """Write JSON through a temporary file, then replace the destination atomically."""
    temp_path = path.with_suffix(path.suffix + ".tmp")
    try:
        temp_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        temp_path.replace(path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def merge_batches(
    existing: List[Dict[str, Any]],
    fetched: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], int, int, int, List[str]]:
    """
    Upsert fetched batches into the existing master by Brewfather _id.

    Existing records not returned by the API are retained. This protects historical
    records if API behaviour, account visibility or status filtering changes later.
    """
    existing_by_id = {
        str(batch.get("_id")): batch
        for batch in existing
        if batch.get("_id") not in (None, "")
    }
    merged_by_id = dict(existing_by_id)

    new_count = 0
    updated_count = 0
    unchanged_count = 0
    changed_ids: List[str] = []

    for fetched_batch in fetched:
        batch_id = fetched_batch.get("_id")
        if batch_id in (None, ""):
            print("Warning: Brewfather returned a batch without _id; record skipped.", file=sys.stderr)
            continue
        batch_id = str(batch_id)
        old_batch = existing_by_id.get(batch_id)

        # Preserve previously downloaded device readings when the current run did
        # not request them.
        if old_batch and "apiReadings" not in fetched_batch and "apiReadings" in old_batch:
            fetched_batch = dict(fetched_batch)
            fetched_batch["apiReadings"] = old_batch["apiReadings"]

        if old_batch is None:
            new_count += 1
            changed_ids.append(batch_id)
        elif canonical_json(old_batch) != canonical_json(fetched_batch):
            updated_count += 1
            changed_ids.append(batch_id)
        else:
            unchanged_count += 1

        merged_by_id[batch_id] = fetched_batch

    # Preserve any malformed historical records without an _id at the end rather
    # than silently deleting them.
    idless_existing = [batch for batch in existing if batch.get("_id") in (None, "")]
    merged = list(merged_by_id.values()) + idless_existing
    merged.sort(key=brew_date_sort_key, reverse=True)
    return merged, new_count, updated_count, unchanged_count, changed_ids


def backup_master(master_path: Path, history_dir: Path, tz_name: str = "Africa/Johannesburg") -> Optional[Path]:
    if not master_path.exists():
        return None
    history_dir.mkdir(parents=True, exist_ok=True)
    stamp = now_in_tz(tz_name).strftime("%Y%m%d_%H%M%S_%f")
    backup_path = history_dir / f"Brewfather_Batches_Master_{stamp}.json"
    shutil.copy2(master_path, backup_path)
    return backup_path


RUN_LOG_HEADER = [
    "run_at",
    "script_version",
    "api_batches_fetched",
    "batches_new",
    "batches_updated",
    "batches_unchanged",
    "batches_historical_retained",
    "batches_master_count",
    "api_recipes_fetched",
    "recipes_new",
    "recipes_updated",
    "recipes_unchanged",
    "recipes_historical_retained",
    "recipes_master_count",
    "master_changed",
    "batches_master_sha256",
    "recipes_master_sha256",
    "backup_files",
    "changed_batch_ids",
    "changed_recipe_ids",
]


def append_run_log(
    log_path: Path,
    run_dt: datetime,
    batches_fetched: int,
    batches_new: int,
    batches_updated: int,
    batches_unchanged: int,
    batches_retained: int,
    batches_count: int,
    recipes_fetched: int,
    recipes_new: int,
    recipes_updated: int,
    recipes_unchanged: int,
    recipes_retained: int,
    recipes_count: int,
    master_changed: bool,
    batches_sha256: str,
    recipes_sha256: str,
    backup_paths: List[Optional[Path]],
    changed_batch_ids: List[str],
    changed_recipe_ids: List[str],
) -> None:
    """Append one audit row per run. The log is never rewritten or pruned."""
    file_exists = log_path.exists()
    with log_path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        if not file_exists:
            writer.writerow(RUN_LOG_HEADER)
        backup_file_names = "; ".join(p.name for p in backup_paths if p)
        writer.writerow(
            [
                run_dt.isoformat(timespec="seconds"),
                SCRIPT_VERSION,
                batches_fetched,
                batches_new,
                batches_updated,
                batches_unchanged,
                batches_retained,
                batches_count,
                recipes_fetched,
                recipes_new,
                recipes_updated,
                recipes_unchanged,
                recipes_retained,
                recipes_count,
                "Y" if master_changed else "N",
                batches_sha256,
                recipes_sha256,
                backup_file_names,
                "; ".join(changed_batch_ids),
                "; ".join(changed_recipe_ids),
            ]
        )


# -----------------------------------------------------------------------------
# Brewfather extraction
# -----------------------------------------------------------------------------


def fetch_brewfather_batches(
    user_id: str,
    api_key: str,
    sleep_between_calls: float = 0.0,
    max_retries: int = 5,
) -> List[Dict[str, Any]]:
    """Fetch all Brewfather batches using v2 pagination with start_after."""
    if requests is None:
        raise RuntimeError("Missing dependency: requests. Install with: pip install requests")

    batches: List[Dict[str, Any]] = []
    start_after: Optional[str] = None
    seen_page_ends: set[str] = set()

    while True:
        params: Dict[str, Any] = {
            "complete": True,
            "limit": MAX_LIMIT,
            "order_by": "_id",
            "order_by_direction": "asc",
        }
        if start_after:
            params["start_after"] = start_after

        response = None
        last_network_error: Optional[str] = None
        for attempt in range(1, max_retries + 1):
            try:
                response = requests.get(
                    f"{API_BASE_URL}/batches",
                    params=params,
                    auth=(user_id, api_key),
                    timeout=60,
                )
            except requests.RequestException as exc:
                last_network_error = str(exc)
                response = None
                delay = min(60, 2 ** attempt)
                print(
                    f"Network error contacting Brewfather: {exc}; retrying in {delay} seconds "
                    f"(attempt {attempt}/{max_retries}).",
                    file=sys.stderr,
                )
                time.sleep(delay)
                continue

            if response.status_code == 429:
                try:
                    retry_after = max(1, int(response.headers.get("Retry-After", "60")))
                except (TypeError, ValueError):
                    retry_after = 60
                print(
                    f"Rate limited by Brewfather; sleeping {retry_after} seconds "
                    f"(attempt {attempt}/{max_retries}).",
                    file=sys.stderr,
                )
                time.sleep(retry_after)
                continue

            if 500 <= response.status_code < 600:
                delay = min(60, 2 ** attempt)
                print(
                    f"Brewfather server error {response.status_code}; retrying in {delay} seconds "
                    f"(attempt {attempt}/{max_retries}).",
                    file=sys.stderr,
                )
                time.sleep(delay)
                continue

            break

        if response is None:
            raise RuntimeError(
                f"No response from Brewfather after {max_retries} attempts. "
                f"Last network error: {last_network_error or 'unknown'}"
            )
        if response.status_code == 429 or response.status_code >= 500:
            raise RuntimeError(
                f"Brewfather API failed after {max_retries} attempts: "
                f"HTTP {response.status_code} {response.text[:500]}"
            )
        try:
            response.raise_for_status()
        except requests.RequestException as exc:
            raise RuntimeError(
                f"Brewfather API request failed: HTTP {response.status_code} {response.text[:1000]}"
            ) from exc

        try:
            page = response.json()
        except ValueError as exc:
            raise RuntimeError(
                f"Brewfather returned invalid JSON: {response.text[:300]}"
            ) from exc
        if not isinstance(page, list):
            raise RuntimeError(
                f"Unexpected Brewfather response. Expected list, got {type(page).__name__}."
            )
        if not page:
            break

        batches.extend(item for item in page if isinstance(item, dict))
        page_end = page[-1].get("_id") if isinstance(page[-1], dict) else None
        if not page_end:
            raise RuntimeError("Pagination failed: last batch in page has no _id.")
        page_end = str(page_end)
        if page_end in seen_page_ends:
            raise RuntimeError("Pagination loop detected: Brewfather returned the same page end twice.")
        seen_page_ends.add(page_end)
        start_after = page_end

        if len(page) < MAX_LIMIT:
            break
        if sleep_between_calls > 0:
            time.sleep(sleep_between_calls)

    return batches

def fetch_batch_readings(
    batch_id: str,
    user_id: str,
    api_key: str,
    max_retries: int = 5,
) -> List[Dict[str, Any]]:
    """Fetch all device readings for one batch with bounded retries."""
    if requests is None:
        return []

    for attempt in range(1, max_retries + 1):
        try:
            response = requests.get(
                f"{API_BASE_URL}/batches/{batch_id}/readings",
                auth=(user_id, api_key),
                timeout=60,
            )
        except requests.RequestException as exc:
            delay = min(60, 2 ** attempt)
            print(
                f"Network error fetching readings for {batch_id}: {exc}; retrying in "
                f"{delay} seconds (attempt {attempt}/{max_retries}).",
                file=sys.stderr,
            )
            time.sleep(delay)
            continue
        if response.status_code == 404:
            return []
        if response.status_code == 429:
            try:
                retry_after = max(1, int(response.headers.get("Retry-After", "60")))
            except (TypeError, ValueError):
                retry_after = 60
            print(
                f"Rate limited while fetching readings for {batch_id}; sleeping "
                f"{retry_after} seconds (attempt {attempt}/{max_retries}).",
                file=sys.stderr,
            )
            time.sleep(retry_after)
            continue
        if 500 <= response.status_code < 600:
            delay = min(60, 2 ** attempt)
            print(
                f"Server error {response.status_code} fetching readings for {batch_id}; "
                f"retrying in {delay} seconds (attempt {attempt}/{max_retries}).",
                file=sys.stderr,
            )
            time.sleep(delay)
            continue
        if response.status_code >= 400:
            print(
                f"Could not fetch readings for batch {batch_id}: HTTP {response.status_code}.",
                file=sys.stderr,
            )
            return []
        try:
            data = response.json()
        except ValueError:
            print(f"Invalid JSON in readings response for batch {batch_id}; skipped.", file=sys.stderr)
            return []
        return [item for item in data if isinstance(item, dict)] if isinstance(data, list) else []

    print(f"Readings for batch {batch_id} skipped after {max_retries} failed attempts.", file=sys.stderr)
    return []


def fetch_brewfather_recipes(
    user_id: str,
    api_key: str,
    sleep_between_calls: float = 0.0,
    max_retries: int = 5,
) -> List[Dict[str, Any]]:
    """Fetch all Brewfather recipes using v2 pagination with start_after."""
    if requests is None:
        raise RuntimeError("Missing dependency: requests. Install with: pip install requests")

    recipes: List[Dict[str, Any]] = []
    start_after: Optional[str] = None
    seen_page_ends: set[str] = set()

    while True:
        params: Dict[str, Any] = {
            "complete": True,
            "limit": MAX_LIMIT,
            "order_by": "_id",
            "order_by_direction": "asc",
        }
        if start_after:
            params["start_after"] = start_after

        response = None
        last_network_error: Optional[str] = None
        for attempt in range(1, max_retries + 1):
            try:
                response = requests.get(
                    f"{API_BASE_URL}/recipes",
                    params=params,
                    auth=(user_id, api_key),
                    timeout=60,
                )
            except requests.RequestException as exc:
                last_network_error = str(exc)
                response = None
                delay = min(60, 2 ** attempt)
                print(
                    f"Network error contacting Brewfather: {exc}; retrying in {delay} seconds "
                    f"(attempt {attempt}/{max_retries}).",
                    file=sys.stderr,
                )
                time.sleep(delay)
                continue

            if response.status_code == 429:
                try:
                    retry_after = max(1, int(response.headers.get("Retry-After", "60")))
                except (TypeError, ValueError):
                    retry_after = 60
                print(
                    f"Rate limited by Brewfather; sleeping {retry_after} seconds "
                    f"(attempt {attempt}/{max_retries}).",
                    file=sys.stderr,
                )
                time.sleep(retry_after)
                continue

            if 500 <= response.status_code < 600:
                delay = min(60, 2 ** attempt)
                print(
                    f"Brewfather server error {response.status_code}; retrying in {delay} seconds "
                    f"(attempt {attempt}/{max_retries}).",
                    file=sys.stderr,
                )
                time.sleep(delay)
                continue

            break

        if response is None:
            raise RuntimeError(
                f"No response from Brewfather after {max_retries} attempts. "
                f"Last network error: {last_network_error or 'unknown'}"
            )
        if response.status_code == 429 or response.status_code >= 500:
            raise RuntimeError(
                f"Brewfather API failed after {max_retries} attempts: "
                f"HTTP {response.status_code} {response.text[:500]}"
            )
        try:
            response.raise_for_status()
        except requests.RequestException as exc:
            raise RuntimeError(
                f"Brewfather API request failed: HTTP {response.status_code} {response.text[:1000]}"
            ) from exc

        try:
            page = response.json()
        except ValueError as exc:
            raise RuntimeError(
                f"Brewfather returned invalid JSON: {response.text[:300]}"
            ) from exc
        if not isinstance(page, list):
            raise RuntimeError(
                f"Unexpected Brewfather response. Expected list, got {type(page).__name__}."
            )
        if not page:
            break

        recipes.extend(item for item in page if isinstance(item, dict))
        page_end = page[-1].get("_id") if isinstance(page[-1], dict) else None
        if not page_end:
            raise RuntimeError("Pagination failed: last recipe in page has no _id.")
        page_end = str(page_end)
        if page_end in seen_page_ends:
            raise RuntimeError("Pagination loop detected: Brewfather returned the same page end twice.")
        seen_page_ends.add(page_end)
        start_after = page_end

        if len(page) < MAX_LIMIT:
            break
        if sleep_between_calls > 0:
            time.sleep(sleep_between_calls)

    return recipes

# -----------------------------------------------------------------------------
# Batch formatting
# -----------------------------------------------------------------------------


def recipe_of(batch: Dict[str, Any]) -> Dict[str, Any]:
    recipe = batch.get("recipe")
    return recipe if isinstance(recipe, dict) else {}


def batch_title(batch: Dict[str, Any]) -> str:
    recipe = recipe_of(batch)
    batch_no = first_value(batch.get("batchNo"), default="")
    name = first_value(recipe.get("name"), batch.get("name"), default="Batch")
    return f"Batch #{batch_no} - {name}" if batch_no != "" else str(name)


def style_name(recipe: Dict[str, Any]) -> str:
    style = recipe.get("style")
    if isinstance(style, dict):
        return str(style.get("name") or "")
    return str(style or "")


def target_actual_rows(batch: Dict[str, Any]) -> List[Tuple[str, str, str]]:
    recipe = recipe_of(batch)
    equipment = recipe.get("equipment") if isinstance(recipe.get("equipment"), dict) else {}
    water = recipe.get("water") if isinstance(recipe.get("water"), dict) else {}

    # Brewfather has no true measured IBU/colour; these are recalculated batch estimates.
    actual_ibu = first_value(batch.get("measuredIbu"), batch.get("estimatedIbu"), default="")
    actual_color = first_value(batch.get("measuredColor"), batch.get("estimatedColor"), default="")

    rows = [
        ("OG", fmt_gravity(recipe.get("og")), fmt_gravity(batch.get("measuredOg"))),
        ("FG", fmt_gravity(recipe.get("fg")), fmt_gravity(batch.get("measuredFg"))),
        ("ABV %", fmt_number(recipe.get("abv"), 2), fmt_number(batch.get("measuredAbv"), 2)),
        ("IBU", fmt_number(recipe.get("ibu"), 1), fmt_number(actual_ibu, 1)),
        ("Colour EBC", fmt_number(recipe.get("color"), 1), fmt_number(actual_color, 1)),
        ("Pre-boil gravity", fmt_gravity(recipe.get("preBoilGravity")), fmt_gravity(batch.get("measuredPreBoilGravity"))),
        ("Post-boil gravity", fmt_gravity(recipe.get("postBoilGravity")), fmt_gravity(batch.get("measuredPostBoilGravity"))),
        ("Mash pH", fmt_number(get_nested(water, "mashPh", default=""), 2), fmt_number(batch.get("measuredMashPh"), 2)),
        ("Batch size L", fmt_number(recipe.get("batchSize"), 2), fmt_number(batch.get("measuredBatchSize"), 2)),
        ("Boil size L", fmt_number(recipe.get("boilSize"), 2), fmt_number(batch.get("measuredBoilSize"), 2)),
        (
            "Kegged / bottling volume L",
            fmt_number(first_value(equipment.get("bottlingVolume"), recipe.get("bottlingVolume"), default=""), 2),
            fmt_number(batch.get("measuredBottlingSize"), 2),
        ),
        ("Brewhouse efficiency %", fmt_number(recipe.get("efficiency"), 2), fmt_number(batch.get("measuredEfficiency"), 2)),
        ("Mash efficiency %", fmt_number(recipe.get("mashEfficiency"), 2), fmt_number(batch.get("measuredMashEfficiency"), 2)),
        ("Measured attenuation %", "", fmt_number(batch.get("measuredAttenuation"), 2)),
    ]
    return rows


def add_target_actual_table(lines: List[str], markdown: bool, batch: Dict[str, Any]) -> None:
    rows = target_actual_rows(batch)
    if markdown:
        lines.append("| Item | Target | Actual |")
        lines.append("|---|---:|---:|")
        for item, target, actual in rows:
            lines.append(f"| {item} | {target} | {actual} |")
    else:
        lines.append(f"{'Item':30} {'Target':>14} {'Actual':>14}")
        lines.append(f"{'-' * 30} {'-' * 14} {'-' * 14}")
        for item, target, actual in rows:
            lines.append(f"{item:30} {target:>14} {actual:>14}")


def format_fermentables(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    items = safe_list(first_value(recipe.get("fermentables"), batch.get("batchFermentables"), default=[]))
    out: List[str] = []
    for item in items:
        name = item.get("name") or ""
        amount = fmt_amount(item.get("amount"), "kg")
        percent = fmt_number(item.get("percentage"), 2)
        supplier = item.get("supplier") or ""
        color = fmt_number(item.get("color"), 3)
        parts = [amount, name]
        line = " ".join([p for p in parts if p]).strip()
        extras = []
        if percent:
            extras.append(f"{percent}%")
        if supplier:
            extras.append(str(supplier))
        if color:
            extras.append(f"{color} EBC")
        if extras:
            line += " (" + ", ".join(extras) + ")"
        if line:
            out.append(f"  - {line}")
    return out or ["  - None recorded"]


def format_hops(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    items = safe_list(first_value(recipe.get("hops"), batch.get("batchHops"), default=[]))
    out: List[str] = []
    for item in items:
        name = item.get("name") or ""
        amount = fmt_amount(item.get("amount"), "g")
        use = item.get("use") or item.get("usage") or ""
        time_value = item.get("time")
        time_text = ""
        if time_value not in (None, ""):
            time_unit = "days" if item.get("timeIsDays") else "min"
            time_text = f", {fmt_number(time_value, 2)} {time_unit}"
        temp = item.get("temp")
        temp_text = f", {fmt_number(temp, 1)} °C" if temp not in (None, "") else ""
        alpha = fmt_number(item.get("alpha"), 2)
        ibu = fmt_number(item.get("ibu"), 1)
        extras = []
        if alpha:
            extras.append(f"{alpha}% AA")
        if ibu:
            extras.append(f"{ibu} IBU")
        head = " ".join(p for p in (amount, str(name)) if p)
        line = f"  - {head}"
        if use:
            line += f" - {use}"
        line += f"{time_text}{temp_text}"
        if extras:
            line += ", " + ", ".join(extras)
        out.append(line)
    return out or ["  - None recorded"]


def format_yeasts(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    items = safe_list(first_value(recipe.get("yeasts"), batch.get("batchYeasts"), default=[]))
    out: List[str] = []
    for item in items:
        name = item.get("name") or ""
        lab = item.get("laboratory") or ""
        product = item.get("productId") or ""
        amount = fmt_amount(item.get("amount"), item.get("unit") or "pkg")
        min_temp = fmt_number(item.get("minTemp"), 1)
        max_temp = fmt_number(item.get("maxTemp"), 1)
        attenuation = fmt_number(item.get("attenuation"), 1)
        notes = clean_text(item.get("userNotes"))

        display = " ".join([p for p in [amount, lab, name] if p]).strip()
        if product:
            display += f" ({product})"
        extras = []
        if min_temp or max_temp:
            extras.append(f"{min_temp}–{max_temp} °C")
        if attenuation:
            extras.append(f"{attenuation}% attenuation")
        if notes:
            extras.append(f"Notes: {notes}")
        if extras:
            display += " - " + "; ".join(extras)
        out.append(f"  - {display}")
    return out or ["  - None recorded"]


def format_miscs(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    items = safe_list(first_value(recipe.get("miscs"), batch.get("batchMiscs"), default=[]))
    out: List[str] = []
    for item in items:
        name = item.get("name") or ""
        amount = fmt_amount(item.get("amount"), item.get("unit") or "")
        use = item.get("use") or ""
        typ = item.get("type") or ""
        time_value = item.get("time")
        time_text = ""
        if time_value not in (None, ""):
            time_unit = "days" if item.get("timeIsDays") else "min"
            time_text = f", {fmt_number(time_value, 2)} {time_unit}"
        details = ", ".join([p for p in [use, typ] if p])
        line = f"  - {amount} {name}"
        if details:
            line += f" - {details}"
        if time_text:
            line += time_text
        out.append(line)
    return out or ["  - None recorded"]


def format_mash(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    mash = recipe.get("mash") if isinstance(recipe.get("mash"), dict) else {}
    out: List[str] = []
    if mash.get("name"):
        out.append(f"  Profile: {mash.get('name')}")
    for step in safe_list(mash.get("steps")):
        name = step.get("name") or step.get("type") or "Step"
        temp = fmt_number(first_value(step.get("displayStepTemp"), step.get("stepTemp"), default=""), 1)
        time_min = fmt_number(step.get("stepTime"), 1)
        ramp = fmt_number(step.get("rampTime"), 1)
        line = f"  - {name}: {temp} °C for {time_min} min" if temp or time_min else f"  - {name}"
        if ramp:
            line += f"; ramp {ramp} min"
        out.append(line)
    return out or ["  - None recorded"]


def format_fermentation(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    ferm = recipe.get("fermentation") if isinstance(recipe.get("fermentation"), dict) else {}
    out: List[str] = []
    if ferm.get("name"):
        out.append(f"  Profile: {ferm.get('name')}")
    for step in safe_list(ferm.get("steps")):
        name = step.get("name") or step.get("type") or "Step"
        temp = fmt_number(first_value(step.get("displayStepTemp"), step.get("stepTemp"), default=""), 1)
        days = fmt_number(step.get("stepTime"), 2)
        pressure = fmt_number(first_value(step.get("displayPressure"), step.get("pressure"), default=""), 1)
        ramp = fmt_number(step.get("ramp"), 2)
        line = f"  - {name}: {temp} °C for {days} days" if temp or days else f"  - {name}"
        if pressure:
            line += f" at {pressure} PSI"
        if ramp:
            line += f"; ramp {ramp}"
        out.append(line)
    return out or ["  - None recorded"]


def format_water(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    water = recipe.get("water") if isinstance(recipe.get("water"), dict) else {}
    profile = first_value(water.get("total"), water.get("mash"), water.get("source"), default={})
    if not isinstance(profile, dict):
        profile = {}

    out: List[str] = []
    name = profile.get("name") or ""
    if name:
        out.append(f"  Profile: {name}")

    ions = {
        "Ca": profile.get("calcium"),
        "Mg": profile.get("magnesium"),
        "Na": profile.get("sodium"),
        "Cl": profile.get("chloride"),
        "SO4": profile.get("sulfate"),
        "HCO3": profile.get("bicarbonate"),
    }
    ion_text = ", ".join(
        f"{key} {fmt_number(value, 2)}"
        for key, value in ions.items()
        if value not in (None, "")
    )
    if ion_text:
        out.append(f"  {ion_text}")

    so_cl = first_value(
        profile.get("soClRatio"),
        get_nested(water, "mash", "soClRatio", default=""),
        get_nested(water, "source", "soClRatio", default=""),
        default="",
    )
    if so_cl not in (None, ""):
        out.append(f"  SO4/Cl ratio: {fmt_number(so_cl, 2)}")
    mash_ph = get_nested(water, "mashPh", default="")
    if mash_ph not in (None, ""):
        out.append(f"  Target mash pH: {fmt_number(mash_ph, 2)}")
    return out or ["  - None recorded"]

def format_notes_log(batch: Dict[str, Any], tz_name: str) -> List[str]:
    log_items: List[Tuple[int, str]] = []

    for note in safe_list(batch.get("notes")):
        ts = note.get("timestamp") or 0
        date_text = fmt_date(ts, tz_name=tz_name, with_time=True)
        status = note.get("status") or ""
        typ = note.get("type") or ""
        text = clean_text(note.get("note"))
        status_part = status
        if typ:
            status_part = f"{status_part} ({typ})" if status_part else typ
        head = " ".join(p for p in (date_text, status_part) if p)
        line = f"  - {head}: {text}" if text else f"  - {head}"
        try:
            sort_key = int(ts)
        except Exception:
            sort_key = 0
        log_items.append((sort_key, line))

    for item in safe_list(batch.get("batchLog")):
        ts = item.get("date") or item.get("timestamp") or item.get("time") or 0
        date_text = fmt_date(ts, tz_name=tz_name, with_time=True)
        typ = item.get("type") or ""
        text = clean_text(first_value(item.get("text"), item.get("note"), item.get("message"), default=""))
        head = " ".join(p for p in (date_text, str(typ)) if p)
        line = f"  - {head}: {text}" if text else f"  - {head}"
        try:
            sort_key = int(ts)
        except Exception:
            sort_key = 0
        log_items.append((sort_key, line))

    log_items.sort(key=lambda x: x[0])
    return [line for _, line in log_items] or ["  - None recorded"]


def format_equipment(batch: Dict[str, Any]) -> List[str]:
    recipe = recipe_of(batch)
    equipment = recipe.get("equipment") if isinstance(recipe.get("equipment"), dict) else {}
    fields = [
        ("Equipment profile", equipment.get("name"), None),
        ("Sparge formula", equipment.get("spargeWaterFormula"), None),
        ("Boil-off per hour L", equipment.get("boilOffPerHr"), 1),
        ("Fermenter loss L", equipment.get("fermenterLoss"), 1),
        ("Trub/chiller loss L", equipment.get("trubChillerLoss"), 1),
    ]
    out: List[str] = []
    for label, value, decimals in fields:
        if value in (None, ""):
            continue
        rendered = fmt_number(value, decimals) if decimals is not None and is_number(value) else str(value)
        out.append(f"{label}: {rendered}")
    return out or ["No equipment details recorded"]

def as_md_bullets(body: List[str]) -> List[str]:
    """
    Convert plain key-value lines to Markdown bullets so consecutive lines do
    not merge into one rendered paragraph. Existing bullets, table rows,
    headings and blank lines pass through unchanged.
    """
    out: List[str] = []
    for line in body:
        stripped = line.strip()
        if not stripped or stripped.startswith(("-", "|", "#")):
            out.append(line)
        else:
            out.append(f"- {stripped}")
    return out


def format_batch(batch: Dict[str, Any], markdown: bool = False, tz_name: str = "Africa/Johannesburg") -> str:
    recipe = recipe_of(batch)
    lines: List[str] = []

    def add_section(title: str, body: List[str], bullets: bool = True) -> None:
        lines.append("")
        section(lines, title, markdown=markdown)
        if markdown and bullets:
            lines.extend(as_md_bullets(body))
        else:
            lines.extend(body)

    title = batch_title(batch)
    if markdown:
        lines.append(f"## {title}")
    else:
        lines.append("=" * 90)
        lines.append(title)
        lines.append("=" * 90)

    overview = [
        f"Status: {batch.get('status') or ''}",
        f"Style: {style_name(recipe)}",
        f"Type: {recipe.get('type') or ''}",
        f"Brewer: {batch.get('brewer') or recipe.get('author') or ''}",
        f"Brew date: {fmt_date(batch.get('brewDate'), tz_name=tz_name)}",
        f"Fermentation start: {fmt_date(batch.get('fermentationStartDate'), tz_name=tz_name)}",
        f"Kegging / bottling date: {fmt_date(batch.get('bottlingDate'), tz_name=tz_name)}",
    ]
    add_section("Overview", overview)

    lines.append("")
    section(lines, "Targets vs actuals", markdown=markdown)
    add_target_actual_table(lines, markdown, batch)

    add_section("Equipment / volumes", format_equipment(batch))
    add_section("Fermentables", format_fermentables(batch))
    add_section("Hops", format_hops(batch))
    add_section("Yeast", format_yeasts(batch))
    add_section("Miscellaneous / water agents / finings", format_miscs(batch))
    add_section("Mash profile", format_mash(batch))
    add_section("Fermentation profile", format_fermentation(batch))
    add_section("Water profile", format_water(batch))

    notes = clean_text(recipe.get("notes"))
    if notes:
        body: List[str] = []
        add_multiline(body, notes)
        add_section("Recipe notes", body, bullets=False)

    batch_notes = clean_text(batch.get("batchNotes"))
    if batch_notes:
        body = []
        add_multiline(body, batch_notes)
        add_section("Batch notes", body, bullets=False)

    taste_rating = batch.get("tasteRating")
    taste_notes = clean_text(batch.get("tasteNotes"))
    if taste_rating not in (None, "") or taste_notes:
        lines.append("")
        section(lines, "Taste", markdown=markdown)
        if taste_rating not in (None, ""):
            rating_line = f"Rating: {taste_rating}"
            lines.append(f"- {rating_line}" if markdown else rating_line)
        if taste_notes:
            add_multiline(lines, taste_notes)

    add_section("Batch log / status notes", format_notes_log(batch, tz_name=tz_name))

    readings = safe_list(batch.get("apiReadings"))
    if readings:
        body = [f"Readings extracted: {len(readings)}"]
        for reading in readings[:20]:
            date_text = fmt_date(first_value(reading.get("time"), reading.get("timestamp"), default=""), tz_name=tz_name, with_time=True)
            sg = fmt_gravity(first_value(reading.get("sg"), reading.get("gravity"), default=""))
            temp = fmt_number(first_value(reading.get("temp"), reading.get("temperature"), default=""), 2)
            device = first_value(reading.get("type"), reading.get("id"), default="")
            head = " ".join(p for p in (date_text, str(device)) if p)
            body.append(f"  - {head}: SG {sg}, temp {temp} °C")
        if len(readings) > 20:
            body.append(f"  - ... {len(readings) - 20} further readings omitted from readable output; full readings are preserved in raw JSON.")
        add_section("Device readings", body)

    lines.append("")
    return "\n".join(lines)


def write_readable_outputs(
    batches: List[Dict[str, Any]],
    txt_path: Path,
    md_path: Path,
    tz_name: str,
    run_timestamp: str,
) -> None:
    txt_lines = [
        f"Brewfather Batches - Clean Readable Master - updated {run_timestamp}",
        "=" * 70,
        "",
        f"Total batches: {len(batches)}",
        "",
    ]
    md_lines = [
        f"# Brewfather Batches - Clean Readable Master - updated {run_timestamp}",
        "",
        f"Total batches: {len(batches)}",
        "",
    ]

    for batch in batches:
        txt_lines.append(format_batch(batch, markdown=False, tz_name=tz_name))
        txt_lines.append("")
        md_lines.append(format_batch(batch, markdown=True, tz_name=tz_name))
        md_lines.append("")

    txt_path.write_text("\n".join(txt_lines), encoding="utf-8")
    md_path.write_text("\n".join(md_lines), encoding="utf-8")

# -----------------------------------------------------------------------------
# Recipe formatting
# -----------------------------------------------------------------------------


def format_recipe(recipe: Dict[str, Any], markdown: bool = False, tz_name: str = "Africa/Johannesburg") -> str:
    """Format a single recipe record as readable text (no measured values, unlike batches)."""
    lines: List[str] = []

    def add_section(title: str, body: List[str], bullets: bool = True) -> None:
        lines.append("")
        section(lines, title, markdown=markdown)
        if markdown and bullets:
            lines.extend(as_md_bullets(body))
        else:
            lines.extend(body)

    name = recipe.get('name', 'Unnamed Recipe')
    recipe_id = recipe.get('_id', '—')
    
    if markdown:
        lines.append(f"## {name}")
    else:
        lines.append("=" * 90)
        lines.append(name)
        lines.append("=" * 90)

    overview = [
        f"Recipe ID: {recipe_id}",
        f"Style: {recipe.get('style', {}).get('name') if isinstance(recipe.get('style'), dict) else recipe.get('style', '')}",
        f"Type: {recipe.get('type', '')}",
        f"Author: {recipe.get('author', '')}",
        f"Created: {fmt_date(recipe.get('_created'), tz_name=tz_name)}",
    ]
    add_section("Overview", overview)

    og = recipe.get('og')
    fg = recipe.get('fg')
    abv = recipe.get('abv')
    ibu = recipe.get('ibu')
    color = recipe.get('color')
    
    metrics = [
        f"OG: {fmt_gravity(og)}",
        f"FG: {fmt_gravity(fg)}",
        f"ABV: {fmt_number(abv, 2)}%",
        f"IBU: {fmt_number(ibu, 1)}",
        f"Color EBC: {fmt_number(color, 1)}",
        f"Batch size L: {fmt_number(recipe.get('batchSize'), 2)}",
        f"Boil size L: {fmt_number(recipe.get('boilSize'), 2)}",
    ]
    add_section("Targets", metrics)

    add_section("Fermentables", format_fermentables({'recipe': recipe}))
    add_section("Hops", format_hops({'recipe': recipe}))
    add_section("Yeast", format_yeasts({'recipe': recipe}))
    add_section("Miscellaneous", format_miscs({'recipe': recipe}))
    add_section("Mash profile", format_mash({'recipe': recipe}))
    add_section("Fermentation profile", format_fermentation({'recipe': recipe}))
    add_section("Water profile", format_water({'recipe': recipe}))

    notes = clean_text(recipe.get('notes'))
    if notes:
        body: List[str] = []
        add_multiline(body, notes)
        add_section("Notes", body, bullets=False)

    lines.append("")
    return "\n".join(lines)


def write_recipe_outputs(
    recipes: List[Dict[str, Any]],
    txt_path: Path,
    md_path: Path,
    tz_name: str,
    run_timestamp: str,
) -> None:
    txt_lines = [
        f"Brewfather Recipes - Clean Readable Master - updated {run_timestamp}",
        "=" * 70,
        "",
        f"Total recipes: {len(recipes)}",
        "",
    ]
    md_lines = [
        f"# Brewfather Recipes - Clean Readable Master - updated {run_timestamp}",
        "",
        f"Total recipes: {len(recipes)}",
        "",
    ]

    for recipe in recipes:
        txt_lines.append(format_recipe(recipe, markdown=False, tz_name=tz_name))
        txt_lines.append("")
        md_lines.append(format_recipe(recipe, markdown=True, tz_name=tz_name))
        md_lines.append("")

    txt_path.write_text("\n".join(txt_lines), encoding="utf-8")
    md_path.write_text("\n".join(md_lines), encoding="utf-8")


def recipe_summary_row(recipe: Dict[str, Any], tz_name: str) -> Dict[str, Any]:
    style = recipe.get('style')
    style_name_val = style.get('name') if isinstance(style, dict) else str(style or '')
    
    return {
        "Recipe ID": recipe.get("_id"),
        "Name": recipe.get("name"),
        "Style": style_name_val,
        "Type": recipe.get("type"),
        "Author": recipe.get("author"),
        "Created": fmt_date(recipe.get("_created"), tz_name=tz_name),
        "Target OG": recipe.get("og"),
        "Target FG": recipe.get("fg"),
        "Target ABV %": recipe.get("abv"),
        "Target IBU": recipe.get("ibu"),
        "Target Colour EBC": recipe.get("color"),
        "Batch Size L": recipe.get("batchSize"),
        "Boil Size L": recipe.get("boilSize"),
        "Brewhouse Efficiency %": recipe.get("efficiency"),
        "Mash Efficiency %": recipe.get("mashEfficiency"),
        "Notes": clean_text(recipe.get("notes")),
    }


def write_recipe_xlsx_summary(recipes: List[Dict[str, Any]], xlsx_path: Path, tz_name: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed; skipping Excel summary. Install with: pip install openpyxl", file=sys.stderr)
        return False

    gravity_columns = {"Target OG", "Target FG"}
    one_decimal_columns = {"Target IBU", "Target Colour EBC"}
    two_decimal_columns = {
        "Target ABV %", "Batch Size L", "Boil Size L",
        "Brewhouse Efficiency %", "Mash Efficiency %",
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "Recipes": [recipe_summary_row(recipe, tz_name) for recipe in recipes],
        "Fermentables": flatten_items(recipes, "Fermentable", "fermentables", "fermentables"),
        "Hops": flatten_items(recipes, "Hop", "hops", "hops"),
        "Yeast": flatten_items(recipes, "Yeast", "yeasts", "yeasts"),
        "Misc": flatten_items(recipes, "Misc", "miscs", "miscs"),
    }

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, data_rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        if not data_rows:
            worksheet.append(["No data"])
            continue

        headers: List[str] = []
        for row in data_rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)

        worksheet.append(headers)
        for row in data_rows:
            worksheet.append([row.get(header, "") for header in headers])

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if sheet_name == "Recipes":
            for column_index, header in enumerate(headers, start=1):
                if header in gravity_columns:
                    number_format = "0.000"
                elif header in one_decimal_columns:
                    number_format = "0.0"
                elif header in two_decimal_columns:
                    number_format = "0.00"
                else:
                    number_format = None
                if number_format:
                    for column_cells in worksheet.iter_rows(
                        min_row=2,
                        max_row=worksheet.max_row,
                        min_col=column_index,
                        max_col=column_index,
                    ):
                        column_cells[0].number_format = number_format

        for column_index, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for cells in worksheet.iter_rows(
                min_row=2,
                max_row=min(worksheet.max_row, 300),
                min_col=column_index,
                max_col=column_index,
            ):
                value = cells[0].value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 10), 55
            )

    workbook.save(xlsx_path)
    return True




def batch_summary_row(batch: Dict[str, Any], tz_name: str) -> Dict[str, Any]:
    recipe = recipe_of(batch)
    equipment = recipe.get("equipment") if isinstance(recipe.get("equipment"), dict) else {}
    water = recipe.get("water") if isinstance(recipe.get("water"), dict) else {}
    actual_ibu = first_value(batch.get("measuredIbu"), batch.get("estimatedIbu"), default=None)
    actual_color = first_value(batch.get("measuredColor"), batch.get("estimatedColor"), default=None)
    return {
        "Batch ID": batch.get("_id"),
        "Batch No": batch.get("batchNo"),
        "Batch Name": first_value(recipe.get("name"), batch.get("name"), default=""),
        "Status": batch.get("status"),
        "Style": style_name(recipe),
        "Type": recipe.get("type"),
        "Brewer": batch.get("brewer") or recipe.get("author"),
        "Brew Date": fmt_date(batch.get("brewDate"), tz_name=tz_name),
        "Fermentation Start": fmt_date(batch.get("fermentationStartDate"), tz_name=tz_name),
        "Kegging/Bottling Date": fmt_date(batch.get("bottlingDate"), tz_name=tz_name),
        "Target OG": recipe.get("og"),
        "Actual OG": batch.get("measuredOg"),
        "Target FG": recipe.get("fg"),
        "Actual FG": batch.get("measuredFg"),
        "Target ABV %": recipe.get("abv"),
        "Actual ABV %": batch.get("measuredAbv"),
        "Target IBU": recipe.get("ibu"),
        "Actual/Estimated IBU": actual_ibu,
        "Target Colour EBC": recipe.get("color"),
        "Actual/Estimated Colour EBC": actual_color,
        "Target Pre-boil Gravity": recipe.get("preBoilGravity"),
        "Actual Pre-boil Gravity": batch.get("measuredPreBoilGravity"),
        "Target Post-boil Gravity": recipe.get("postBoilGravity"),
        "Actual Post-boil Gravity": batch.get("measuredPostBoilGravity"),
        "Target Mash pH": get_nested(water, "mashPh", default=None),
        "Actual Mash pH": batch.get("measuredMashPh"),
        "Target Batch Size L": recipe.get("batchSize"),
        "Actual Batch Size L": batch.get("measuredBatchSize"),
        "Target Boil Size L": recipe.get("boilSize"),
        "Actual Boil Size L": batch.get("measuredBoilSize"),
        "Target Packaging Volume L": first_value(
            equipment.get("bottlingVolume"), recipe.get("bottlingVolume"), default=None
        ),
        "Actual Packaging Volume L": batch.get("measuredBottlingSize"),
        "Target Brewhouse Efficiency %": recipe.get("efficiency"),
        "Actual Brewhouse Efficiency %": batch.get("measuredEfficiency"),
        "Target Mash Efficiency %": recipe.get("mashEfficiency"),
        "Actual Mash Efficiency %": batch.get("measuredMashEfficiency"),
        "Measured Attenuation %": batch.get("measuredAttenuation"),
        "Taste Rating": batch.get("tasteRating"),
        "Recipe Notes": clean_text(recipe.get("notes")),
        "Batch Notes": clean_text(batch.get("batchNotes")),
        "Taste Notes": clean_text(batch.get("tasteNotes")),
    }

def flatten_items(batches: List[Dict[str, Any]], item_name: str, recipe_key: str, batch_key: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for batch in batches:
        recipe = recipe_of(batch)
        items = safe_list(first_value(recipe.get(recipe_key), batch.get(batch_key), default=[]))
        for item in items:
            row = {
                "Batch ID": batch.get("_id"),
                "Batch No": batch.get("batchNo"),
                "Batch Name": first_value(recipe.get("name"), batch.get("name"), default=""),
                "Item Type": item_name,
            }
            for key, value in item.items():
                if isinstance(value, (dict, list)):
                    row[key] = json.dumps(value, ensure_ascii=False)
                else:
                    row[key] = value
            rows.append(row)
    return rows


def flatten_logs(batches: List[Dict[str, Any]], tz_name: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for batch in batches:
        recipe = recipe_of(batch)
        base = {
            "Batch ID": batch.get("_id"),
            "Batch No": batch.get("batchNo"),
            "Batch Name": first_value(recipe.get("name"), batch.get("name"), default=""),
        }
        for note in safe_list(batch.get("notes")):
            rows.append({
                **base,
                "Source": "notes",
                "Date": fmt_date(note.get("timestamp"), tz_name=tz_name, with_time=True),
                "Status": note.get("status"),
                "Type": note.get("type"),
                "Note": clean_text(note.get("note")),
            })
        for item in safe_list(batch.get("batchLog")):
            timestamp = first_value(item.get("date"), item.get("timestamp"), item.get("time"), default="")
            rows.append({
                **base,
                "Source": "batchLog",
                "Date": fmt_date(timestamp, tz_name=tz_name, with_time=True),
                "Status": item.get("status"),
                "Type": item.get("type"),
                "Note": clean_text(first_value(item.get("text"), item.get("note"), item.get("message"), default="")),
            })
    return rows

def write_xlsx_summary(batches: List[Dict[str, Any]], xlsx_path: Path, tz_name: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed; skipping Excel summary. Install with: pip install openpyxl", file=sys.stderr)
        return False

    gravity_columns = {
        "Target OG", "Actual OG", "Target FG", "Actual FG",
        "Target Pre-boil Gravity", "Actual Pre-boil Gravity",
        "Target Post-boil Gravity", "Actual Post-boil Gravity",
    }
    one_decimal_columns = {
        "Target IBU", "Actual/Estimated IBU",
        "Target Colour EBC", "Actual/Estimated Colour EBC",
    }
    two_decimal_columns = {
        "Target ABV %", "Actual ABV %", "Target Mash pH", "Actual Mash pH",
        "Target Batch Size L", "Actual Batch Size L", "Target Boil Size L",
        "Actual Boil Size L", "Target Packaging Volume L", "Actual Packaging Volume L",
        "Target Brewhouse Efficiency %", "Actual Brewhouse Efficiency %",
        "Target Mash Efficiency %", "Actual Mash Efficiency %", "Measured Attenuation %",
    }

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {
        "Batches": [batch_summary_row(batch, tz_name) for batch in batches],
        "Fermentables": flatten_items(batches, "Fermentable", "fermentables", "batchFermentables"),
        "Hops": flatten_items(batches, "Hop", "hops", "batchHops"),
        "Yeast": flatten_items(batches, "Yeast", "yeasts", "batchYeasts"),
        "Misc": flatten_items(batches, "Misc", "miscs", "batchMiscs"),
        "Logs": flatten_logs(batches, tz_name),
    }

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, data_rows in sheets.items():
        worksheet = workbook.create_sheet(sheet_name)
        if not data_rows:
            worksheet.append(["No data"])
            continue

        headers: List[str] = []
        for row in data_rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)

        worksheet.append(headers)
        for row in data_rows:
            worksheet.append([row.get(header, "") for header in headers])

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if sheet_name == "Batches":
            for column_index, header in enumerate(headers, start=1):
                if header in gravity_columns:
                    number_format = "0.000"
                elif header in one_decimal_columns:
                    number_format = "0.0"
                elif header in two_decimal_columns:
                    number_format = "0.00"
                elif header in {"Batch No", "Taste Rating"}:
                    number_format = "0"
                else:
                    number_format = None
                if number_format:
                    for column_cells in worksheet.iter_rows(
                        min_row=2,
                        max_row=worksheet.max_row,
                        min_col=column_index,
                        max_col=column_index,
                    ):
                        column_cells[0].number_format = number_format

        for column_index, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for cells in worksheet.iter_rows(
                min_row=2,
                max_row=min(worksheet.max_row, 300),
                min_col=column_index,
                max_col=column_index,
            ):
                value = cells[0].value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(
                max(max_length + 2, 10), 55
            )

        # Keep long notes readable rather than expanding row height uncontrollably.
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 80:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    workbook.save(xlsx_path)
    return True

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    default_out_dir = Path(__file__).resolve().parent / "brewfather_exports"
    parser = argparse.ArgumentParser(
        description="Update a local Brewfather batch master and regenerate clean outputs."
    )
    parser.add_argument(
        "--user-id",
        default=os.getenv("BREWFATHER_USER_ID"),
        help="Brewfather User ID. Prefer BREWFATHER_USER_ID environment variable.",
    )
    parser.add_argument(
        "--api-key",
        default=os.getenv("BREWFATHER_API_KEY"),
        help="Brewfather API key. Prefer BREWFATHER_API_KEY environment variable.",
    )
    parser.add_argument(
        "--out-dir",
        default=str(default_out_dir),
        help=f"Output directory. Default: {default_out_dir}",
    )
    parser.add_argument(
        "--timezone",
        default="Africa/Johannesburg",
        help="Timezone for date conversion. Default: Africa/Johannesburg.",
    )
    parser.add_argument(
        "--include-readings",
        action="store_true",
        help="Fetch device readings for new or changed batches. Adds API calls.",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.0,
        help="Seconds to sleep between API calls.",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Do not generate the Excel summary.",
    )
    parser.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not archive the previous master JSON before a changed update.",
    )
    parser.add_argument(
        "--no-run-log",
        action="store_true",
        help="Do not append a row to the run audit log (Brewfather_Runs_Log.csv).",
    )
    return parser.parse_args()

def main() -> int:
    args = parse_args()
    if not args.user_id or not args.api_key:
        print(
            "Missing Brewfather credentials. Set BREWFATHER_USER_ID and "
            "BREWFATHER_API_KEY, or pass --user-id and --api-key.",
            file=sys.stderr,
        )
        return 2

    out_dir = Path(args.out_dir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    
    # Batch paths
    batches_master_path = out_dir / "Brewfather_Batches_Master.json"
    batches_txt_path = out_dir / "Brewfather_Batches_Readable_Clean.txt"
    batches_md_path = out_dir / "Brewfather_Batches_Readable_Clean.md"
    batches_xlsx_path = out_dir / "Brewfather_Batches_Summary.xlsx"
    
    # Recipe paths
    recipes_master_path = out_dir / "Brewfather_Recipes_Master.json"
    recipes_txt_path = out_dir / "Brewfather_Recipes_Readable_Clean.txt"
    recipes_md_path = out_dir / "Brewfather_Recipes_Readable_Clean.md"
    recipes_xlsx_path = out_dir / "Brewfather_Recipes_Summary.xlsx"
    
    run_log_path = out_dir / "Brewfather_Runs_Log.csv"
    history_dir = out_dir / "history"

    # Fetch batches and recipes
    batches_existing = load_batch_list(batches_master_path)
    batches_fetched = fetch_brewfather_batches(
        user_id=args.user_id,
        api_key=args.api_key,
        sleep_between_calls=max(0.0, args.sleep),
    )
    
    recipes_existing = load_batch_list(recipes_master_path)
    recipes_fetched = fetch_brewfather_recipes(
        user_id=args.user_id,
        api_key=args.api_key,
        sleep_between_calls=max(0.0, args.sleep),
    )

    # Determine batch changes before optional readings calls.
    batches_existing_by_id = {
        str(batch.get("_id")): batch
        for batch in batches_existing
        if batch.get("_id") not in (None, "")
    }
    batches_changed_or_new_ids = {
        str(batch.get("_id"))
        for batch in batches_fetched
        if batch.get("_id") not in (None, "")
        and (
            str(batch.get("_id")) not in batches_existing_by_id
            or canonical_json(without_api_readings(batches_existing_by_id[str(batch.get("_id"))]))
            != canonical_json(without_api_readings(batch))
        )
    }

    # Fetch device readings for changed batches if requested
    if args.include_readings:
        for batch in batches_fetched:
            batch_id = batch.get("_id")
            if batch_id in (None, ""):
                continue
            batch_id = str(batch_id)
            old_batch = batches_existing_by_id.get(batch_id)
            needs_readings = (
                batch_id in batches_changed_or_new_ids
                or old_batch is None
                or "apiReadings" not in old_batch
            )
            if needs_readings:
                readings = fetch_batch_readings(batch_id, args.user_id, args.api_key)
                if readings:
                    batch["apiReadings"] = readings
                elif old_batch and "apiReadings" in old_batch:
                    batch["apiReadings"] = old_batch["apiReadings"]
                if args.sleep > 0:
                    time.sleep(args.sleep)

    # Merge batches and recipes
    batches_merged, batches_new, batches_updated, batches_unchanged, batches_changed_ids = merge_batches(batches_existing, batches_fetched)
    recipes_merged, recipes_new, recipes_updated, recipes_unchanged, recipes_changed_ids = merge_batches(recipes_existing, recipes_fetched)
    
    batches_master_changed = canonical_json(batches_existing) != canonical_json(batches_merged)
    recipes_master_changed = canonical_json(recipes_existing) != canonical_json(recipes_merged)
    master_changed = batches_master_changed or recipes_master_changed

    # Backup and write masters
    backup_paths: List[Optional[Path]] = []
    
    if batches_master_changed and batches_master_path.exists() and not args.no_backup:
        backup_paths.append(backup_master(batches_master_path, history_dir, tz_name=args.timezone))
    if batches_master_changed or not batches_master_path.exists():
        write_json_atomic(batches_master_path, batches_merged)
    
    if recipes_master_changed and recipes_master_path.exists() and not args.no_backup:
        backup_paths.append(backup_master(recipes_master_path, history_dir, tz_name=args.timezone))
    if recipes_master_changed or not recipes_master_path.exists():
        write_json_atomic(recipes_master_path, recipes_merged)

    run_dt = now_in_tz(args.timezone)
    run_timestamp = run_dt.strftime("%Y-%m-%d %H:%M")
    
    # Write outputs
    write_readable_outputs(
        batches_merged,
        batches_txt_path,
        batches_md_path,
        tz_name=args.timezone,
        run_timestamp=run_timestamp,
    )
    
    write_recipe_outputs(
        recipes_merged,
        recipes_txt_path,
        recipes_md_path,
        tz_name=args.timezone,
        run_timestamp=run_timestamp,
    )

    batches_excel_created = False
    recipes_excel_created = False
    if not args.no_excel:
        batches_excel_created = write_xlsx_summary(batches_merged, batches_xlsx_path, tz_name=args.timezone)
        recipes_excel_created = write_recipe_xlsx_summary(recipes_merged, recipes_xlsx_path, tz_name=args.timezone)

    batches_retained = max(0, len(batches_merged) - len({str(b.get("_id")) for b in batches_fetched if b.get("_id")}))
    recipes_retained = max(0, len(recipes_merged) - len({str(r.get("_id")) for r in recipes_fetched if r.get("_id")}))

    batches_sha256 = hashlib.sha256(canonical_json(batches_merged).encode("utf-8")).hexdigest()
    recipes_sha256 = hashlib.sha256(canonical_json(recipes_merged).encode("utf-8")).hexdigest()
    
    if not args.no_run_log:
        append_run_log(
            run_log_path,
            run_dt=run_dt,
            batches_fetched=len(batches_fetched),
            batches_new=batches_new,
            batches_updated=batches_updated,
            batches_unchanged=batches_unchanged,
            batches_retained=batches_retained,
            batches_count=len(batches_merged),
            recipes_fetched=len(recipes_fetched),
            recipes_new=recipes_new,
            recipes_updated=recipes_updated,
            recipes_unchanged=recipes_unchanged,
            recipes_retained=recipes_retained,
            recipes_count=len(recipes_merged),
            master_changed=master_changed,
            batches_sha256=batches_sha256,
            recipes_sha256=recipes_sha256,
            backup_paths=backup_paths,
            changed_batch_ids=batches_changed_ids,
            changed_recipe_ids=recipes_changed_ids,
        )

    print("Brewfather update complete")
    print("")
    print("BATCHES:")
    print(f"  API fetched:              {len(batches_fetched)}")
    print(f"  New added:                {batches_new}")
    print(f"  Existing updated:         {batches_updated}")
    print(f"  Unchanged:                {batches_unchanged}")
    print(f"  Historical retained:      {batches_retained}")
    print(f"  Master count:             {len(batches_merged)}")
    print(f"  Master SHA-256:           {batches_sha256}")
    print("")
    print("RECIPES:")
    print(f"  API fetched:              {len(recipes_fetched)}")
    print(f"  New added:                {recipes_new}")
    print(f"  Existing updated:         {recipes_updated}")
    print(f"  Unchanged:                {recipes_unchanged}")
    print(f"  Historical retained:      {recipes_retained}")
    print(f"  Master count:             {len(recipes_merged)}")
    print(f"  Master SHA-256:           {recipes_sha256}")
    print("")
    print("FILES:")
    print(f"  Batches JSON:             {batches_master_path}")
    print(f"  Batches readable:         {batches_txt_path}, {batches_md_path}")
    print(f"  Batches Excel:            {batches_xlsx_path if batches_excel_created else 'not created'}")
    print(f"  Recipes JSON:             {recipes_master_path}")
    print(f"  Recipes readable:         {recipes_txt_path}, {recipes_md_path}")
    print(f"  Recipes Excel:            {recipes_xlsx_path if recipes_excel_created else 'not created'}")
    print(f"  Run log:                  {run_log_path if not args.no_run_log else 'disabled'}")
    if backup_paths:
        for backup in backup_paths:
            if backup:
                print(f"  Backup:                   {backup}")
    if batches_changed_ids or recipes_changed_ids:
        if batches_changed_ids:
            print(f"  Changed batch IDs:        {', '.join(batches_changed_ids)}")
        if recipes_changed_ids:
            print(f"  Changed recipe IDs:       {', '.join(recipes_changed_ids)}")
    else:
        print("  No changes detected; clean outputs were refreshed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
